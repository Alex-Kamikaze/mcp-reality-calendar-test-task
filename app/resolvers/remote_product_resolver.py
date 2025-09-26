import os
from typing import List
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook
from models.base_resolver import DataResolver
from models.data_models import ProductFromExcelModel
from exceptions.remote_resolver_exceptions import (
    ResolvingException,
    FileNotFoundException,
)


class GoogleDataResolver(DataResolver):
    def __init__(self, creds_path: str):
        self.creds_path = creds_path

        self.SCOPES = ["https://www.googleapis.com/auth/drive"]
        creds = None
        if os.path.exists(self.creds_path):
            creds = Credentials.from_authorized_user_file(
                self.creds_path, scopes=self.SCOPES
            )

        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    "./credentials.json", self.SCOPES
                )
                creds = flow.run_local_server(port=0, prompt="consent")
            with open("token.json", "w") as token:
                token.write(creds.to_json())

        self.service = build("drive", "v3", credentials=creds)

    def _get_files_from_drive(self) -> List[str]:
        """
        Загружает список доступных файлов с Google Drive

        :return: Список файлов для скачивания
        """
        try:
            results = (
                self.service.files()
                .list(pageSize=100, fields="nextPageToken, files(id, name, mimeType)")
                .execute()
            )
            items = results.get("files", [])
            return items
        except HttpError:
            raise ResolvingException()

    def _download_file_from_drive(self, file_id: str):
        """
        Скачивает файл с Google Drive

        :param file_id: ID файла на диске
        """

        try:
            request = self.service.files().get_media(fileId=file_id)
            with open("./data.xlsx", "wb") as fh:
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while not done:
                    _, done = downloader.next_chunk()
        except HttpError:
            raise ResolvingException()

    def get_data_from_xlsx(self, file_name: str) -> List[ProductFromExcelModel]:
        """
        Получает данные из файлов Excel, лежащих на Google-диске

        :param token_path: Путь к токену для авторизации на Google Drive API
        :return: Возвращает список инструментами с их именами и ссылками на описания
        """
        try:
            files = self._get_files_from_drive()
            filtred_file = list(
                filter(lambda file: file.get("name") == file_name, files)
            )
            if not filtred_file:
                raise FileNotFoundException()

            filtred_file = filtred_file[0]
            self._download_file_from_drive(filtred_file.get("id"))

            wb = load_workbook("./data.xlsx")
            ws = wb.active
            os.remove("./data.xlsx")
            return [
                ProductFromExcelModel(name=str(row[0]), link=str(row[1]))
                for row in ws.iter_rows(values_only=True)
            ]
        except HttpError:
            raise ResolvingException()